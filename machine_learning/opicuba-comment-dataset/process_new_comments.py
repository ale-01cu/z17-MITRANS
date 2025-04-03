import pandas as pd
import uuid
from datetime import datetime
import os
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

"""
Pregunta -> Aclaración / Duda / Consulta / Solicitud de información
Ofensa -> Insulto / Discriminación / Acoso / Violencia verbal
Opinión -> Crítica constructiva / Comentario neutral / Elogio
Denuncia -> Incumplimiento de normas / Fraude / Situación ilegal
Queja -> Reclamo / Insatisfacción / Frustración
Sugerencia -> Propuesta de mejora / Solicitud específica
"""

def post_already_exists(df, post_text, post_date):
    """Check if a post with same text and date already exists"""
    if df.empty:
        return False
    return ((df['post_text'] == post_text) & 
            (df['post_date'] == post_date)).any()

def process_post(comments_file, post_text, post_date, output_excel_file, profile, fuente):
    """
    Processes a CSV file with comments and adds them to a general dataset in Excel.
    Now includes profile and source fields.
    """
    
    # Convert to absolute paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    comments_file = os.path.join(script_dir, comments_file)
    output_excel_file = os.path.join(script_dir, output_excel_file)
    
    # Read comments from CSV
    try:
        comments_df = pd.read_csv(comments_file)
        if 'comments' not in comments_df.columns:
            raise ValueError("El archivo CSV debe contener una columna llamada 'comments'")
    except Exception as e:
        print(f"Error leyendo CSV: {e}")
        return
    
    # Check if output file exists
    if os.path.exists(output_excel_file):
        try:
            existing_df = pd.read_excel(output_excel_file, engine='openpyxl')
            
            # Check for duplicate post
            if post_already_exists(existing_df, post_text, post_date):
                print(f"⚠️ El post ya existe en el dataset. No se agregarán datos duplicados.")
                print(f"📌 Post: {post_text[:50]}...")
                print(f"📅 Fecha: {post_date}")
                print(f"👤 Perfil: {profile}")
                print(f"🌐 Fuente: {fuente}")
                return
                
        except Exception as e:
            print(f"Error leyendo Excel: {e}")
            existing_df = pd.DataFrame()
    else:
        # Create the output directory if it doesn't exist
        os.makedirs(os.path.dirname(output_excel_file), exist_ok=True)
        print(f"📁 Creando nuevo archivo Excel: {output_excel_file}")
        existing_df = pd.DataFrame()
    
    # Generate unique IDs only if we're adding new data
    post_id = str(uuid.uuid4())[:8]
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Create DataFrame for new comments with additional fields
    data = {
        'post_id': [post_id] * len(comments_df),
        'post_date': [post_date] * len(comments_df),
        'post_text': [post_text] * len(comments_df),
        'profile': [profile] * len(comments_df),
        'fuente': [fuente] * len(comments_df),
        'comment_id': [f"comm_{post_id}_{i}" for i in range(len(comments_df))],
        'comment_text': comments_df['comments'].tolist(),
        'label': [""] * len(comments_df),
        'processing_date': [current_timestamp] * len(comments_df)
    }
    new_df = pd.DataFrame(data)
    
    # Combine with existing data
    final_df = pd.concat([existing_df, new_df], ignore_index=True)
    
    # Save to Excel with validation
    try:
        # First save with pandas
        final_df.to_excel(output_excel_file, index=False, engine='openpyxl')
        
        # Add data validation for label column
        wb = load_workbook(output_excel_file)
        ws = wb.active
        
        dv = DataValidation(type="list", 
                          formula1='"pregunta,ofensa,opinion,denuncia,queja,sugerencia"', 
                          allow_blank=True)
        
        # Find label column position dynamically
        label_col = None
        for cell in ws[1]:  # Check header row
            if cell.value == "label":
                label_col = cell.column_letter
                break
                
        if label_col:
            dv.add(f'{label_col}2:{label_col}{len(final_df)+1}')
            ws.add_data_validation(dv)
        
        wb.save(output_excel_file)
        print(f"✅ Datos guardados exitosamente en {output_excel_file}")
        print(f"📊 Estadísticas:")
        print(f"- 📝 Post agregado: {post_text[:50]}...")
        print(f"- 👤 Perfil: {profile}")
        print(f"- 🌐 Fuente: {fuente}")
        print(f"- 💬 Comentarios agregados: {len(new_df)}")
        print(f"- 📅 Fecha del post: {post_date}")
        
    except Exception as e:
        print(f"Error guardando Excel: {e}")

# Example usage
if __name__ == "__main__":
    # Configuration (modify these values)
    comments_file = "./input/facebook-cubadebate(viernes, 28 de marzo de 2025 a las 922).csv"  # CSV with 'comments' column
    post_text = """Autoridades médicas esclarecen caso de niño cubano
Publicaciones en redes digitales han reflejado el caso de un #NiñoCubano de diez años de edad que fue atendido en diferentes instituciones de salud de la capital, como parte del diagnóstico y tratamiento a las enfermedades que padece. En aras de explicar los procedimientos que se siguen en las instituciones en #Cuba ante casos de tal complejidad médica y, en consonancia con la probada ética y sensibilidad del gremio de la salud cubano, comparecen autoridades del Hospital Pediátrico Juan Manuel Márquez, de los Institutos de Hematología e Inmunología, de Neurología y Neurocirugía, así como del Ministerio de Salud Pública.."""
    profile = "Cubadebate"               # Nombre del perfil/autor
    fuente = "Facebook"                        # Plataforma de origen
    post_date = "2025-03-20 9:22:00"  # Format: YYYY-MM-DD HH:MM:SS
    dataset_excel = "./output/opicuba-dataset.xlsx"  # Excel that accumulates all data
    
    # Process the post
    process_post(
        comments_file=comments_file,
        post_text=post_text,
        post_date=post_date,
        output_excel_file=dataset_excel,
        profile=profile,
        fuente=fuente
    )